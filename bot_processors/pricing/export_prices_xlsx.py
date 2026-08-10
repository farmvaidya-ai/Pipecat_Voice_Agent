"""One-off export of last_known_prices.json to an .xlsx file for manual
review in Excel — sorting/filtering by state, district, or price is easier
there than reading the raw JSON. This does not change what the bot itself
reads/writes; the fact_market_prices table (db/schema.sql) is the actual
runtime store now — last_known_prices.json is only kept for this export
and enrich_prices_with_geo.py.

Usage: python -m bot_processors.pricing.export_prices_xlsx [output_path.xlsx] [--date=DD-MM-YYYY]

--date filters to rows whose arrival_date matches exactly — since
last_known_prices.json is a rolling "latest known per key" cache, not a
dated history, this only surfaces whatever rows currently still carry that
arrival_date (i.e. haven't since been overwritten by a newer scrape).
Useful for spot-checking a specific day's rows against the live site.
"""

import json
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from bot_processors.pricing.price_shared import _LAST_KNOWN_KEY_SEP, _LAST_KNOWN_PATH

_COLUMNS = [
    "state", "district", "commodity", "market",
    "arrival_date", "modal_per_kg", "min_per_kg", "max_per_kg", "stale",
    "arrival_qty", "arrival_unit",
]


def export_to_xlsx(output_path: Path, arrival_date: str | None = None) -> int:
    entries = json.loads(_LAST_KNOWN_PATH.read_text(encoding="utf-8"))
    if arrival_date is not None:
        entries = {k: v for k, v in entries.items() if v.get("arrival_date") == arrival_date}

    wb = Workbook()
    ws = wb.active
    ws.title = "last_known_prices"
    ws.append(_COLUMNS)

    for key, row in sorted(entries.items()):
        state, district, crop_keyword = key.split(_LAST_KNOWN_KEY_SEP)
        ws.append([row.get(col) for col in _COLUMNS])

    for i, col in enumerate(_COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = max(len(col) + 2, 14)
    ws.freeze_panes = "A2"

    wb.save(output_path)
    return len(entries)


if __name__ == "__main__":
    date_filter = None
    positional = []
    for arg in sys.argv[1:]:
        if arg.startswith("--date="):
            date_filter = arg.split("=", 1)[1]
        else:
            positional.append(arg)

    out = Path(positional[0]) if positional else _LAST_KNOWN_PATH.with_suffix(".xlsx")
    if not _LAST_KNOWN_PATH.exists():
        print(f"No {_LAST_KNOWN_PATH} found — nothing to export.")
        sys.exit(1)
    count = export_to_xlsx(out, arrival_date=date_filter)
    print(f"Exported {count} row(s) to {out}")
