"""Exports nationwide_prices.jsonl (see
bot_processors.agmarknet_scraper.scrape_nationwide) to an .xlsx file — one
row per state/district/market/commodity combo scraped, exactly as the site
reported it (raw units, not converted to per-kg like the live bot's cache),
so it's a faithful source for a future database import.

Usage: python -m bot_processors.export_nationwide_xlsx [output_path.xlsx]

Safe to run against a still-in-progress scrape_nationwide() run — it only
reads whatever complete lines exist in the file at that moment.
"""

import json
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from bot_processors.agmarknet_scraper import _NATIONWIDE_OUTPUT_PATH


def export_to_xlsx(output_path: Path) -> int:
    # Column set isn't fixed ahead of time: a Price-only run has 12 columns,
    # a Both-mode run has 15 (adds Arrival Quantity/Unit/Date). Read every
    # row first so the header reflects whichever mode actually produced this
    # file, in the order first seen.
    records: list[dict] = []
    columns: list[str] = []
    with _NATIONWIDE_OUTPUT_PATH.open("r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            record = json.loads(line)
            for row in record["rows"]:
                records.append(row)
                for col in row:
                    if col not in columns:
                        columns.append(col)

    wb = Workbook()
    ws = wb.active
    ws.title = "nationwide_prices"
    ws.append(columns)

    for row in records:
        ws.append([row.get(col) for col in columns])

    for i, col in enumerate(columns, start=1):
        ws.column_dimensions[get_column_letter(i)].width = max(len(col) + 2, 14)
    ws.freeze_panes = "A2"

    wb.save(output_path)
    return len(records)


if __name__ == "__main__":
    out = Path(sys.argv[1]) if len(sys.argv) > 1 else _NATIONWIDE_OUTPUT_PATH.with_suffix(".xlsx")
    if not _NATIONWIDE_OUTPUT_PATH.exists():
        print(f"No {_NATIONWIDE_OUTPUT_PATH} found — nothing to export.")
        sys.exit(1)
    count = export_to_xlsx(out)
    print(f"Exported {count} row(s) to {out}")
